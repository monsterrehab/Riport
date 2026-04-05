import streamlit as st
import pandas as pd
import io
from datetime import date
import openpyxl

# Oldal beállításai
st.set_page_config(page_title="Heti Riport", page_icon="📝", layout="centered")
st.title("📱 Heti Riport Generáló")

# Felejtő memória a termékeknek
if 'termekek' not in st.session_state:
    st.session_state.termekek = []

# --- 1. ALAPADATOK ---
st.header("1. Általános adatok")
col1, col2 = st.columns(2)
with col1:
    nev = st.text_input("Név", value="Danyi Róbert")
    uzlet = st.text_input("Üzlet", value="MM Westend")
with col2:
    st.info("Ezek a neved és az üzlet, ami minden rögzített sorba bekerül.")

# --- 2. NAPI MUNKA RÖGZÍTÉSE ---
st.header("2. Napi eladások rögzítése")
napi_col1, napi_col2 = st.columns(2)
with napi_col1:
    datum = st.date_input("Dátum", value=date.today())
    het = datum.isocalendar()[1]
    magyar_napok = ["hétfő", "kedd", "szerda", "csütörtök", "péntek", "szombat", "vasárnap"]
    nap = magyar_napok[datum.weekday()]
    st.write(f"📅 **{het}. hét, {nap}**")
with napi_col2:
    oraszam = st.number_input("Ledolgozott óraszám", min_value=1, value=8)
    megszolitott = st.number_input("Megszólított vásárlók száma", min_value=0, value=0)

# Terméklista kiolvasása a sablonból
termek_lista = []
try:
    df_fokusz = pd.read_excel('sablon.xlsx', sheet_name='Fókusz', header=None)
    termek_lista = df_fokusz[0].dropna().tolist()
    if "Típus" in termek_lista: 
        termek_lista.remove("Típus")
except Exception:
    termek_lista = ["Kérlek töltsd fel a sablon.xlsx fájlt a GitHubra!"]

if "No sales" not in termek_lista:
    termek_lista.insert(0, "No sales")

t_col1, t_col2 = st.columns([2, 1])
with t_col1:
    kivalasztott_tipus = st.selectbox("Termék típusa", termek_lista)
    alap_darab = 0 if kivalasztott_tipus == "No sales" else 1
    darab = st.number_input("Darabszám", min_value=0, value=alap_darab, step=1)
with t_col2:
    ar = st.number_input("Polcár (Ft)", min_value=0, step=1000)
    
if st.button("➕ Hozzáadás a heti listához"):
    st.session_state.termekek.append({
        "Név": nev,
        "Üzlet": uzlet,
        "Dátum": datum.strftime("%Y-%m-%d"),
        "Hét": het,
        "Nap": nap,
        "Ledolgozott óraszám": oraszam,
        "Megszólított vásárlók száma": megszolitott,
        "Típus": kivalasztott_tipus,
        "Ár": ar,
        "Darab": darab
    })
    st.success(f"{darab}x {kivalasztott_tipus} sikeresen rögzítve a listára ({nap})!")

if st.session_state.termekek:
    df_mutato = pd.DataFrame(st.session_state.termekek)[["Nap", "Típus", "Darab", "Ár"]]
    st.dataframe(df_mutato)
    if st.button("🗑️ Teljes lista törlése"):
        st.session_state.termekek = []
        st.rerun()

# --- 3. HETI SZÖVEGES RÉSZ ---
st.header("3. Heti szöveges összefoglaló")
st.info("💡 Ezeket a mezőket elég egyszer kitölteni a hét végén, az exportálás előtt!")
hiany = st.text_area("Milyen termék hiányzik / mit rendeltetnél?", value="-")
konkurencia = st.text_area("Konkurencia (új brand, promóter, akciók, legkeresettebb termék)", value="-")
kihelyezes = st.text_area("Kihelyezés (hiányzó DEMO, planogram, működnek-e)", value="Minden rendben")
komment = st.text_area("Komment (tapasztalatok, észrevételek)", placeholder="Heti összefoglaló az áruházzal, vásárlókkal kapcsolatban...")

# --- 4. EXPORTÁLÁS ---
st.header("4. Exportálás")

if not st.session_state.termekek:
    st.info("ℹ️ A letöltéshez kérlek rögzíts legalább egy terméket vagy a 'No sales' opciót!")
else:
    hibak = []
    for termek in st.session_state.termekek:
        if termek["Típus"] != "No sales":
            if termek["Ár"] == 0:
                hibak.append(f"A(z) {termek['Típus']} termék polcára nem lehet 0 Ft ({termek['Nap']})!")
            if termek["Darab"] == 0:
                hibak.append(f"A(z) {termek['Típus']} termék darabszáma nem lehet 0 ({termek['Nap']})!")

    if hibak:
        st.warning("A fájl letöltéséhez kérlek javítsd a következőket:")
        for hiba in hibak:
            st.error(f"❌ {hiba}")
    else:
        st.success("✅ Minden adat rendben, a riport készen áll a letöltésre!")
        
        try:
            # Eredeti sablon betöltése
            wb = openpyxl.load_workbook('sablon.xlsx')
            if 'Riport' in wb.sheetnames:
                ws = wb['Riport']
            else:
                ws = wb.active
                
            # TISZTÍTÓ ELV: Minden régi adat kiradírozása a 2. sortól lefelé
            for r in range(2, ws.max_row + 1):
                for c in range(1, 15):
                    ws.cell(row=r, column=c).value = None
            
            # A 2. sortól kezdjük a tiszta lapra történő kitöltést
            kezdo_sor = 2
            
            for i, termek in enumerate(st.session_state.termekek):
                sor = kezdo_sor + i
                ws.cell(row=sor, column=1).value = termek["Név"]
                ws.cell(row=sor, column=2).value = termek["Üzlet"]
                ws.cell(row=sor, column=3).value = termek["Dátum"]
                ws.cell(row=sor, column=4).value = termek["Hét"]
                ws.cell(row=sor, column=5).value = termek["Nap"]
                ws.cell(row=sor, column=6).value = termek["Ledolgozott óraszám"]
                ws.cell(row=sor, column=7).value = termek["Megszólított vásárlók száma"]
                ws.cell(row=sor, column=8).value = termek["Darab"]
                ws.cell(row=sor, column=9).value = termek["Ár"]
                ws.cell(row=sor, column=10).value = termek["Típus"]
                
                # A heti szövegeket csak a legelső sorba írjuk be
                if i == 0:
                    ws.cell(row=sor, column=11).value = hiany
                    ws.cell(row=sor, column=12).value = konkurencia
                    ws.cell(row=sor, column=13).value = kihelyezes
                    ws.cell(row=sor, column=14).value = komment
            
            buffer = io.BytesIO()
            wb.save(buffer)
            
            formazott_nev = nev.replace(" ", "_")
            utolso_het = st.session_state.termekek[-1]["Hét"]
            fajlnev = f"{formazott_nev}_W{utolso_het}_riport.xlsx"
            
            st.download_button(
                label="📥 Végleges Excel letöltése",
                data=buffer.getvalue(),
                file_name=fajlnev,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        except Exception as e:
            st.error(f"Hiba történt a sablon megnyitásakor. Biztosan feltöltötted a 'sablon.xlsx' fájlt? Részletek: {e}")
